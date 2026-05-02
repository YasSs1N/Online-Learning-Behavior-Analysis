"""
data_prep_oulad.py
==================
OULAD preprocessing for Task 1 (course-level FP-Growth + PageRank).

Reads:
  data/raw/oulad/studentInfo.csv
  data/raw/oulad/studentRegistration.csv
  data/raw/oulad/courses.csv
  data/processed/oulad_module_mapping.xlsx   (code_module -> module_name + domain)

Writes (data/processed/):
  module_mapping.csv        flat mapping (code_module, module_name, domain, ...)
  enrolments_clean.csv      one row per (student, module) -- multi-enrolments preserved
  transactions.csv          one row per student, JSON list of module_names
  course_meta.csv           per-module aggregates (students, pass-rate, avg credits, ...)
  course_graph_edges.csv    co-enrolment edges: (source, target, weight=#students who took both)
"""

from __future__ import annotations

from itertools import combinations
from pathlib import Path

import json
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "oulad"
PROCESSED = ROOT / "data" / "processed"
PROCESSED.mkdir(parents=True, exist_ok=True)


def load_mapping() -> pd.DataFrame:
    """Read the user-curated module-name mapping from the xlsx."""
    path = PROCESSED / "oulad_module_mapping.xlsx"
    wb = load_workbook(path)
    ws = wb["Module Mapping"] if "Module Mapping" in wb.sheetnames else wb.active
    rows = []
    for row in ws.iter_rows(min_row=5, max_row=11, values_only=True):
        code, name, domain, subject, *_ = row
        if code:
            rows.append({"code_module": code, "module_name": name,
                         "domain": domain, "subject_area": subject})
    return pd.DataFrame(rows)


def main() -> None:
    print("[oulad] Loading raw OULAD")
    info = pd.read_csv(RAW / "studentInfo.csv")
    reg = pd.read_csv(RAW / "studentRegistration.csv")
    courses = pd.read_csv(RAW / "courses.csv")

    mapping = load_mapping()
    mapping.to_csv(PROCESSED / "module_mapping.csv", index=False)
    print(f"[oulad] Mapping has {len(mapping)} modules")
    print(mapping.to_string(index=False))

    # Merge readable name onto every row
    info_named = info.merge(mapping, on="code_module", how="left")

    # ---- Per (student, module) enrolments ---- #
    enrolments = (info_named[["id_student", "code_module", "module_name", "domain",
                              "code_presentation", "final_result", "studied_credits"]]
                  .drop_duplicates(subset=["id_student", "code_module"]))
    enrolments.to_csv(PROCESSED / "enrolments_clean.csv", index=False)
    print(f"[oulad] Unique (student, module) pairs: {len(enrolments):,}")

    # ---- Transactions: one row per student, list of module_names ---- #
    transactions = (enrolments.groupby("id_student")["module_name"]
                    .apply(lambda s: sorted(set(s))).reset_index(name="modules"))
    transactions["n_modules"] = transactions["modules"].apply(len)
    transactions["modules"] = transactions["modules"].apply(json.dumps)
    transactions.to_csv(PROCESSED / "transactions.csv", index=False)
    n_multi = (transactions["n_modules"] >= 2).sum()
    print(f"[oulad] Total students: {len(transactions):,}  |  multi-module (>=2): {n_multi:,}")
    print(f"[oulad] Modules per student distribution:\n{transactions['n_modules'].value_counts().sort_index()}")

    # ---- Per-module aggregates ---- #
    course_meta = (info_named.groupby(["code_module", "module_name", "domain", "subject_area"])
                   .agg(students=("id_student", "nunique"),
                        presentations=("code_presentation", "nunique"),
                        avg_credits=("studied_credits", "mean"),
                        pct_pass=("final_result", lambda s: (s == "Pass").mean()),
                        pct_distinction=("final_result", lambda s: (s == "Distinction").mean()),
                        pct_fail=("final_result", lambda s: (s == "Fail").mean()),
                        pct_withdrawn=("final_result", lambda s: (s == "Withdrawn").mean()))
                   .round(3).reset_index())
    course_meta = course_meta.merge(courses.groupby("code_module")["module_presentation_length"]
                                    .mean().round(0).reset_index().rename(
                                        columns={"module_presentation_length": "avg_length_days"}),
                                    on="code_module")
    course_meta.to_csv(PROCESSED / "course_meta.csv", index=False)
    print(f"\n[oulad] Course meta:\n{course_meta.to_string(index=False)}")

    # ---- Co-enrolment graph edges ---- #
    print("[oulad] Building co-enrolment graph edges")
    edges = []
    for modules_json in transactions["modules"]:
        modules = json.loads(modules_json)
        if len(modules) < 2:
            continue
        for a, b in combinations(sorted(set(modules)), 2):
            edges.append((a, b))
    edges_df = pd.DataFrame(edges, columns=["source", "target"])
    if not edges_df.empty:
        edges_df = edges_df.groupby(["source", "target"]).size().reset_index(name="weight")
        edges_df = edges_df.sort_values("weight", ascending=False).reset_index(drop=True)
    edges_df.to_csv(PROCESSED / "course_graph_edges.csv", index=False)
    print(f"[oulad] Edges: {len(edges_df)}")
    print(edges_df.head(15).to_string(index=False))
    print("[oulad] Done.")


if __name__ == "__main__":
    main()
